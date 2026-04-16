"""Build one rank job from test/testd1.xlsx: local files + downloaded URLs.

Excel layout (as in testd1.xlsx):
  Col A: reference transcript (Thai); long cells update ``current_target`` for following rows.
  Col C: voice / label name
  Col D: audio URL or local filename (e.g. ``299.mp3``, ``botnoi voice.mp3``)

URLs are saved under ``--download-dir``. All resolved files are uploaded in row order
(duplicates by same URL or same resolved path are skipped).

Requires: ``pip install openpyxl`` (not listed in repo root requirements).

Example::

    python test/run_rank_from_testd1_xlsx.py \\
      --base-url https://since-supporting-edwards-comments.trycloudflare.com \\
      --local-root \"D:\\\\Downloads\\\\泰语\"
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import time
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

import requests

AUDIO_EXTS = (".mp3", ".wav", ".m4a", ".flac", ".ogg", ".aac", ".webm")


def _configure_stdout_utf8() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        try:
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass


def _safe_filename(name: str) -> str:
    keep: list[str] = []
    for ch in name.strip():
        if ch.isalnum() or ch in {"-", "_", "."}:
            keep.append(ch)
        else:
            keep.append("_")
    out = "".join(keep).strip("._") or "audio"
    return out[:160]


def _classify_audio_cell(val: Any) -> tuple[str, str] | None:
    """Return (kind, value) where kind is 'url' or 'file', or None if not an audio ref."""
    if not isinstance(val, str):
        return None
    s = val.strip()
    if not s:
        return None
    if s.startswith("http://") or s.startswith("https://"):
        return ("url", s)
    low = s.lower()
    if any(low.endswith(ext) for ext in AUDIO_EXTS):
        return ("file", s)
    return None


def _pick_target_from_cell(a_val: Any) -> str | None:
    if not isinstance(a_val, str):
        return None
    t = a_val.strip()
    if len(t) < 50:
        return None
    if t.lower().startswith("http://") or t.lower().startswith("https://"):
        return None
    return t


def _resolve_local_audio(name: str, local_root: Path) -> Path | None:
    name = name.strip()
    direct = local_root / name
    if direct.is_file():
        return direct
    hits = [p for p in local_root.rglob(name) if p.is_file()]
    if not hits:
        return None
    hits.sort(key=lambda p: (len(p.parts), str(p)))
    return hits[0]


def _download_url(url: str, dest: Path) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    headers = {"User-Agent": "SpeechJudge-test-run_rank_from_testd1_xlsx/1.0"}
    with requests.get(url, stream=True, timeout=(30, 180), headers=headers, allow_redirects=True) as resp:
        resp.raise_for_status()
        with dest.open("wb") as f:
            for chunk in resp.iter_content(chunk_size=1024 * 1024):
                if chunk:
                    f.write(chunk)


def _submit_multipart(
    base_url: str,
    target_text: str,
    paths: list[tuple[str, Path]],
    submit_timeout: int,
) -> str:
    multipart: list[tuple[str, tuple[str, Any, str]]] = []
    opened: list[Any] = []
    try:
        for upload_name, p in paths:
            f = p.open("rb")
            opened.append(f)
            multipart.append(("audio_files", (upload_name, f, "application/octet-stream")))
        resp = requests.post(
            f"{base_url.rstrip('/')}/jobs/rank",
            data={"target_text": target_text},
            files=multipart,
            timeout=submit_timeout,
        )
    finally:
        for f in opened:
            f.close()
    if resp.status_code >= 400:
        raise RuntimeError(f"POST /jobs/rank failed: HTTP {resp.status_code}, body={resp.text}")
    payload = resp.json()
    job_id = payload.get("job_id")
    if not job_id:
        raise RuntimeError(f"missing job_id: {json.dumps(payload, ensure_ascii=False)}")
    return str(job_id)


def _poll_job(base_url: str, job_id: str, timeout_seconds: int, interval: float) -> dict[str, Any]:
    deadline = time.time() + timeout_seconds
    url = f"{base_url.rstrip('/')}/jobs/{job_id}"
    last: dict[str, Any] = {}
    while time.time() < deadline:
        resp = requests.get(url, timeout=60)
        resp.raise_for_status()
        last = resp.json()
        status = str(last.get("status", ""))
        phase = str(last.get("phase", ""))
        progress = float(last.get("progress", 0.0) or 0.0)
        print(f"[poll] status={status} phase={phase} progress={progress:.4f}")
        if status in {"succeeded", "failed"}:
            return last
        time.sleep(interval)
    raise TimeoutError(f"job did not finish in {timeout_seconds}s; last keys={list(last)}")


def collect_inputs_from_xlsx(
    xlsx: Path,
    *,
    local_root: Path,
    download_dir: Path,
    dry_run: bool,
) -> tuple[str, list[tuple[str, Path]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("openpyxl is required: pip install openpyxl") from exc

    wb = load_workbook(xlsx, read_only=False, data_only=True)
    ws = wb.active
    current_target: str | None = None
    seen_urls: set[str] = set()
    seen_paths: set[str] = set()
    out_paths: list[tuple[str, Path]] = []
    row_errors: list[str] = []

    for r in range(2, (ws.max_row or 0) + 1):
        a_val = ws.cell(r, 1).value
        new_t = _pick_target_from_cell(a_val)
        if new_t:
            current_target = new_t

        voice_raw = ws.cell(r, 3).value
        audio_raw = ws.cell(r, 4).value
        voice = str(voice_raw).strip() if voice_raw is not None else f"row{r}"
        classified = _classify_audio_cell(audio_raw)
        if not classified:
            continue

        kind, ref = classified
        if current_target is None:
            row_errors.append(f"row {r}: audio present but no target_text yet (column A)")
            continue

        if kind == "url":
            if ref in seen_urls:
                continue
            seen_urls.add(ref)
            parsed = urlparse(ref)
            base = Path(parsed.path).name or "audio"
            if not re.search(r"\.[a-zA-Z0-9]{1,8}$", base):
                base = f"{base}.mp3"
            dest = download_dir / f"{len(out_paths):04d}_{_safe_filename(voice)}_{_safe_filename(base)}"
            if not dry_run:
                print(f"[download] {ref} -> {dest}")
                _download_url(ref, dest)
            else:
                print(f"[dry-run] would download -> {dest}")
            key = str(dest.resolve()) if dest.exists() else str(dest)
            if key in seen_paths:
                continue
            seen_paths.add(key)
            upload_name = f"{len(out_paths):04d}_{_safe_filename(voice)}{dest.suffix.lower()}"
            out_paths.append((upload_name, dest))
        else:
            loc = _resolve_local_audio(ref, local_root)
            if loc is None:
                row_errors.append(f"row {r}: local file not found for {ref!r} under {local_root}")
                continue
            key = str(loc.resolve())
            if key in seen_paths:
                continue
            seen_paths.add(key)
            upload_name = f"{len(out_paths):04d}_{_safe_filename(voice)}{loc.suffix.lower()}"
            out_paths.append((upload_name, loc))

    wb.close()

    if row_errors:
        print("[warn] row issues:")
        for e in row_errors[:30]:
            print(f"  {e}")
        if len(row_errors) > 30:
            print(f"  ... and {len(row_errors) - 30} more")

    if current_target is None:
        raise RuntimeError("Could not determine target_text from column A (need a long text cell).")
    if not out_paths:
        raise RuntimeError("No audio rows resolved (check column D and --local-root).")
    return current_target, out_paths


def main() -> int:
    _configure_stdout_utf8()
    parser = argparse.ArgumentParser(description="Download URLs from testd1.xlsx + upload all audios for one rank job.")
    parser.add_argument("--xlsx", type=Path, default=Path(__file__).resolve().parent / "testd1.xlsx")
    parser.add_argument("--local-root", type=Path, default=Path(r"D:\Downloads\泰语"))
    parser.add_argument(
        "--download-dir",
        type=Path,
        default=Path(r"D:\Downloads\泰语\_from_excel_urls"),
        help="Where to store files downloaded from URLs in the sheet",
    )
    parser.add_argument("--base-url", required=True, help="Rank API base URL (tunnel or localhost)")
    parser.add_argument("--dry-run", action="store_true", help="Only parse + download; do not POST /jobs/rank")
    parser.add_argument(
        "--submit-timeout",
        type=int,
        default=1800,
        help="POST timeout seconds (many/large files may need more)",
    )
    parser.add_argument(
        "--job-timeout",
        type=int,
        default=28800,
        help="Max seconds to wait for job (bubble sort: ~n*(n-1)/2 comparisons)",
    )
    parser.add_argument("--poll-interval", type=float, default=5.0)
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(__file__).resolve().parent / "testd1_rank_result.json",
        help="Write final job JSON here (UTF-8)",
    )
    args = parser.parse_args()

    args.download_dir.mkdir(parents=True, exist_ok=True)

    target_text, paths = collect_inputs_from_xlsx(
        args.xlsx,
        local_root=args.local_root,
        download_dir=args.download_dir,
        dry_run=args.dry_run,
    )
    n = len(paths)
    est = n * (n - 1) // 2 if n > 1 else 0
    print(f"[info] target_text length={len(target_text)} chars")
    print(f"[info] uploads={n} (~{est} pairwise comparisons if full bubble sort)")
    for name, p in paths[:5]:
        print(f"  {name} <- {p}")
    if len(paths) > 5:
        print(f"  ... ({len(paths) - 5} more)")

    if args.dry_run:
        print("[dry-run] skipping POST")
        return 0

    job_id = _submit_multipart(args.base_url, target_text, paths, args.submit_timeout)
    print(f"[info] job_id={job_id}")
    result = _poll_job(args.base_url, job_id, args.job_timeout, args.poll_interval)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[info] wrote {args.output}")
    if str(result.get("status")) != "succeeded":
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
