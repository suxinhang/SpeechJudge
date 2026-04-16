"""Shared helpers for testd1.xlsx → local files + manifest (no HTTP rank API)."""

from __future__ import annotations

import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

import requests

AUDIO_EXTS = (".mp3", ".wav", ".m4a", ".flac", ".ogg", ".aac", ".webm")


def safe_filename(name: str) -> str:
    keep: list[str] = []
    for ch in name.strip():
        if ch.isalnum() or ch in {"-", "_", "."}:
            keep.append(ch)
        else:
            keep.append("_")
    out = "".join(keep).strip("._") or "audio"
    return out[:160]


def classify_audio_cell(val: Any) -> tuple[str, str] | None:
    if not isinstance(val, str):
        return None
    s = val.strip()
    if not s:
        return None
    if s.startswith("http://") or s.startswith("https://"):
        return ("url", s)
    low = s.lower()
    if any(low.endswith(ext) for ext in AUDIO_EXTS):
        return ("file", s)
    return None


def pick_target_from_cell(a_val: Any) -> str | None:
    if not isinstance(a_val, str):
        return None
    t = a_val.strip()
    if len(t) < 50:
        return None
    if t.lower().startswith("http://") or t.lower().startswith("https://"):
        return None
    return t


def resolve_local_audio(name: str, local_root: Path) -> Path | None:
    name = name.strip()
    direct = local_root / name
    if direct.is_file():
        return direct
    hits = [p for p in local_root.rglob(name) if p.is_file()]
    if not hits:
        return None
    hits.sort(key=lambda p: (len(p.parts), str(p)))
    return hits[0]


def download_url(url: str, dest: Path) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    headers = {"User-Agent": "SpeechJudge-test-testd1_inputs/1.0"}
    with requests.get(url, stream=True, timeout=(30, 180), headers=headers, allow_redirects=True) as resp:
        resp.raise_for_status()
        with dest.open("wb") as f:
            for chunk in resp.iter_content(chunk_size=1024 * 1024):
                if chunk:
                    f.write(chunk)


@dataclass
class MaterializeResult:
    target_text: str
    """(upload_name, path, voice, source url|local)"""
    items: list[tuple[str, Path, str, str]]
    row_errors: list[str]


def materialize_from_xlsx(
    xlsx: Path,
    *,
    local_root: Path,
    download_dir: Path,
    dry_run: bool,
    do_download_urls: bool,
) -> MaterializeResult:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("openpyxl is required: pip install openpyxl") from exc

    download_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(xlsx, read_only=False, data_only=True)
    ws = wb.active
    current_target: str | None = None
    seen_urls: set[str] = set()
    seen_paths: set[str] = set()
    out: list[tuple[str, Path, str, str]] = []
    row_errors: list[str] = []

    for r in range(2, (ws.max_row or 0) + 1):
        a_val = ws.cell(r, 1).value
        new_t = pick_target_from_cell(a_val)
        if new_t:
            current_target = new_t

        voice_raw = ws.cell(r, 3).value
        audio_raw = ws.cell(r, 4).value
        voice = str(voice_raw).strip() if voice_raw is not None else f"row{r}"
        classified = classify_audio_cell(audio_raw)
        if not classified:
            continue

        kind, ref = classified
        if current_target is None:
            row_errors.append(f"row {r}: audio present but no target_text yet (column A)")
            continue

        if kind == "url":
            if ref in seen_urls:
                continue
            seen_urls.add(ref)
            parsed = urlparse(ref)
            base = Path(parsed.path).name or "audio"
            if not re.search(r"\.[a-zA-Z0-9]{1,8}$", base):
                base = f"{base}.mp3"
            dest = download_dir / f"{len(out):04d}_{safe_filename(voice)}_{safe_filename(base)}"
            if do_download_urls and not dry_run:
                print(f"[download] {ref} -> {dest}")
                download_url(ref, dest)
            elif dry_run:
                print(f"[dry-run] would download -> {dest}")
            key = str(dest.resolve()) if dest.exists() else str(dest)
            if key in seen_paths:
                continue
            seen_paths.add(key)
            upload_name = f"{len(out):04d}_{safe_filename(voice)}{dest.suffix.lower()}"
            out.append((upload_name, dest, voice, ref))
        else:
            loc = resolve_local_audio(ref, local_root)
            if loc is None:
                row_errors.append(f"row {r}: local file not found for {ref!r} under {local_root}")
                continue
            key = str(loc.resolve())
            if key in seen_paths:
                continue
            seen_paths.add(key)
            upload_name = f"{len(out):04d}_{safe_filename(voice)}{loc.suffix.lower()}"
            out.append((upload_name, loc, voice, ref))

    wb.close()

    if current_target is None:
        raise RuntimeError("Could not determine target_text from column A (need a long text cell).")
    if not out:
        raise RuntimeError("No audio rows resolved (check column D and --local-root).")

    return MaterializeResult(target_text=current_target, items=out, row_errors=row_errors)


def verify_files_non_empty(items: list[tuple[str, Path, str, str]]) -> list[str]:
    bad: list[str] = []
    for upload_name, p, _voice, _src in items:
        try:
            if not p.is_file():
                bad.append(f"missing: {upload_name} ({p})")
            elif p.stat().st_size <= 0:
                bad.append(f"empty file: {upload_name} ({p})")
        except OSError as exc:
            bad.append(f"{upload_name} ({p}): {exc}")
    return bad


def manifest_dict(result: MaterializeResult) -> dict[str, Any]:
    n = len(result.items)
    est = n * (n - 1) // 2 if n > 1 else 0
    return {
        "version": 1,
        "target_text": result.target_text,
        "n_items": n,
        "estimated_pairwise_comparisons": est,
        "sort_note": "Server uses bubble sort over items (see rank_worker._bubble_sort_sync).",
        "items": [
            {
                "upload_name": up,
                "path": str(p.resolve()),
                "voice": voice,
                "source": src,
            }
            for up, p, voice, src in result.items
        ],
    }


def write_manifest(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def load_manifest(path: Path) -> dict[str, Any]:
    raw = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(raw, dict) or raw.get("version") != 1:
        raise ValueError("manifest must be a JSON object with version: 1")
    items = raw.get("items")
    if not isinstance(items, list) or not items:
        raise ValueError("manifest.items must be a non-empty list")
    return raw
